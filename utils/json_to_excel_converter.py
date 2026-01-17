from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Dict, Any, List
import json


def convert_json_to_excel(json_data: Dict[str, Any], output_path: str) -> str:
    """
    Convert JSON questionnaire to surveyCTO ready Excel format.
    
    SurveyCTO Excel format typically has these columns:
    - type: Question type (text, integer, select_one, select_multiple, etc.)
    - name: Question variable name
    - label: Question text
    - required: Whether the question is required
    - constraint: Validation rules
    - choice_list: For select questions, reference to choices
    """
    
    wb = Workbook()
    
    # Survey sheet
    survey_sheet = wb.active
    survey_sheet.title = "survey"
    
    # Choices sheet
    choices_sheet = wb.create_sheet("choices")
    
    # Setup survey sheet headers
    survey_headers = ['type', 'name', 'label', 'required', 'constraint', 'relevant', 'appearance', 'hint']
    for col, header in enumerate(survey_headers, start=1):
        cell = survey_sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Setup choices sheet headers
    choices_headers = ['list_name', 'name', 'label']
    for col, header in enumerate(choices_headers, start=1):
        cell = choices_sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Process questions
    questions = json_data.get('questions', [])
    survey_row = 2
    choices_row = 2
    
    for question in questions:
        q_id = question.get('id', '')
        q_text = question.get('label', question.get('question', ''))
        q_type = question.get('type', 'text')
        q_required = question.get('required', False)
        choices = question.get('choices', [])
        
        # Adjust type for surveyCTO
        surveycto_type = q_type
        if choices and q_type in ['select_one', 'select_multiple']:
            list_name = f"{q_id}_list"
            surveycto_type = f"{q_type} {list_name}"
            
            # Add choices to choices sheet
            for idx, choice in enumerate(choices, start=1):
                choices_sheet.cell(row=choices_row, column=1, value=list_name)
                choices_sheet.cell(row=choices_row, column=2, value=f"choice_{idx}")
                choices_sheet.cell(row=choices_row, column=3, value=choice)
                choices_row += 1
        
        # Add question to survey sheet
        survey_sheet.cell(row=survey_row, column=1, value=surveycto_type)
        survey_sheet.cell(row=survey_row, column=2, value=q_id)
        survey_sheet.cell(row=survey_row, column=3, value=q_text)
        survey_sheet.cell(row=survey_row, column=4, value='yes' if q_required else 'no')
        
        # Add hint if available
        if 'description' in question or 'hint' in question:
            hint = question.get('hint', question.get('description', ''))
            survey_sheet.cell(row=survey_row, column=8, value=hint)
        
        survey_row += 1
    
    # Auto-adjust column widths
    for sheet in [survey_sheet, choices_sheet]:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(output_path)
    return output_path
